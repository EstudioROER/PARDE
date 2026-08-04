import openpyxl
import os
import datetime

def append_batch3():
    excel_path = "docs/Inmobiliarias_Interesadas_Peñablanca_PARDE.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    
    nuevos = [
        {
            "empresa": "Constructora e Inmobiliaria Ingevec",
            "tipo": "Inmobiliaria",
            "ubicacion": "Santiago / Presencia Nacional",
            "contacto": "Desarrollo Inmobiliario",
            "email_1": "hola@ingevecinmobiliaria.cl",
            "email_2": "contacto@ingevec.cl",
            "email_3": "",
            "telefono": "No especificado",
            "web": "ingevecinmobiliaria.cl",
            "asunto": "Oportunidad Desarrollo Urbano Peñablanca (Cabida 3D 5.400m2)",
        },
        {
            "empresa": "Inmobiliaria y Constructora Moller y Pérez-Cotapos",
            "tipo": "Inmobiliaria",
            "ubicacion": "Santiago / Regiones",
            "contacto": "Gerencia Comercial",
            "email_1": "mpccentral@moller.cl",
            "email_2": "mpcventas@moller.cl",
            "email_3": "",
            "telefono": "+56 2 2412 2200",
            "web": "mpc.cl",
            "asunto": "Terreno 5.400m2 Peñablanca para Desarrollo Inmobiliario MPC",
        },
        {
            "empresa": "Inmobiliaria Siena",
            "tipo": "Inmobiliaria",
            "ubicacion": "Santiago / V Región",
            "contacto": "Área de Nuevos Negocios",
            "email_1": "ventas@siena.cl",
            "email_2": "contacto@siena.cl",
            "email_3": "",
            "telefono": "+56 2 2571 7500",
            "web": "siena.cl",
            "asunto": "Prospecto Terreno Peñablanca - 5.400m2 para Desarrollo Habitacional Siena",
        },
        {
            "empresa": "Inmobiliaria Actual",
            "tipo": "Inmobiliaria",
            "ubicacion": "Santiago / Regiones",
            "contacto": "Desarrollo / Nuevos Proyectos",
            "email_1": "ventas@actual.cl",
            "email_2": "serviciocliente@actual.cl",
            "email_3": "",
            "telefono": "+56 2 3210 0800",
            "web": "actual.cl",
            "asunto": "Oportunidad Adquisición Suelo Urbano Peñablanca (Cabida 3D Lista)",
        },
        {
            "empresa": "Constructora EBCO",
            "tipo": "Constructora / Inmobiliaria",
            "ubicacion": "Presencia Nacional",
            "contacto": "Área Inmobiliaria / Omcorp",
            "email_1": "contacto@ebco-omcorp.cl",
            "email_2": "contacto@ebco.cl",
            "email_3": "",
            "telefono": "+56 2 2464 4700",
            "web": "ebco.cl",
            "asunto": "Terreno Estratégico 5.400m2 V Región - Oportunidad Desarrollo EBCO",
        }
    ]

    firma = (
        "\n\nAtentamente,\n\n"
        "Sebastián Muñoz Vera & Adolfo Henríquez Carvajal | Parde Arquitectos\n"
        "Arquitectura Técnica & Viabilidad Comercial\n"
        "Especialistas en desbloqueo de activos inmobiliarios y aceleración de ventas.\n"
        "+56 9 5019 6861 / +56 9 7108 9393 | www.pardearquitectos.com"
    )

    mensaje_base = (
        "Estimado equipo de {empresa},\n\n"
        "Mi nombre es Sebastián Muñoz y represento a los propietarios de un terreno estratégico de 5.400 m² ubicado en el sector de Peñablanca, un área de alta expansión y plusvalía en la V Región.\n\n"
        "Considerando su trayectoria en el desarrollo de proyectos inmobiliarios, este paño representa una excelente oportunidad. Para mitigar el riesgo normativo y facilitar su análisis de pre-factibilidad, en PARDE ARQUITECTOS hemos desarrollado una auditoría y una cabida arquitectónica 3D completa.\n\n"
        "Pueden explorar el modelo 3D y la normativa directamente en este enlace interactivo:\n"
        "https://estudioroer.github.io/terreno-penablanca/\n\n"
        "Nos gustaría saber si estarían interesados en evaluar esta oportunidad para su pipeline. Si es así, podemos agendar una breve llamada o enviarles los informes técnicos detallados.\n\n"
        "Quedo a su disposición."
    )

    # 1. Hoja 1
    sheet1 = wb["1. Inmobiliarias Objetivo"]
    max_row_1 = sheet1.max_row
    for idx, c in enumerate(nuevos):
        row = max_row_1 + 1 + idx
        sheet1.cell(row=row, column=1, value=c['empresa'])
        sheet1.cell(row=row, column=2, value=c['ubicacion'])
        sheet1.cell(row=row, column=3, value=c['contacto'])
        sheet1.cell(row=row, column=4, value=c['email_1'])
        sheet1.cell(row=row, column=5, value=c['email_2'])
        sheet1.cell(row=row, column=6, value=c['email_3'])
        sheet1.cell(row=row, column=7, value=c['telefono'])
        sheet1.cell(row=row, column=8, value=c['web'])
        sheet1.cell(row=row, column=9, value="Nueva investigación online")

    # 2. Hoja 3
    sheet3 = wb["3. Correos Personalizados"]
    max_row_3 = sheet3.max_row
    for idx, c in enumerate(nuevos):
        row = max_row_3 + 1 + idx
        sheet3.cell(row=row, column=1, value=max_row_3 - 3 + idx)
        sheet3.cell(row=row, column=2, value=c['tipo'])
        sheet3.cell(row=row, column=3, value=c['empresa'])
        sheet3.cell(row=row, column=4, value=c['email_1'])
        sheet3.cell(row=row, column=5, value=c['email_2'])
        sheet3.cell(row=row, column=6, value=c['email_3'])
        sheet3.cell(row=row, column=7, value=c['telefono'])
        sheet3.cell(row=row, column=8, value=c['web'])
        sheet3.cell(row=row, column=9, value=c['asunto'])
        
        texto = mensaje_base.format(empresa=c['empresa'])
        sheet3.cell(row=row, column=10, value=texto + firma)

    # 3. Hoja 4
    sheet4 = wb["4. Pipeline Seguimiento"]
    max_row_4 = sheet4.max_row
    today_str = datetime.datetime.now().strftime("%d/%m/%Y")
    for idx, c in enumerate(nuevos):
        row = max_row_4 + 1 + idx
        sheet4.cell(row=row, column=1, value=today_str)
        sheet4.cell(row=row, column=2, value=c['empresa'])
        sheet4.cell(row=row, column=3, value=c['contacto'])
        sheet4.cell(row=row, column=4, value="Por definir")
        sheet4.cell(row=row, column=5, value=c['telefono'])
        sheet4.cell(row=row, column=6, value=c['email_1'])
        sheet4.cell(row=row, column=7, value=c['email_2'])
        sheet4.cell(row=row, column=8, value=c['email_3'])
        sheet4.cell(row=row, column=9, value="Email Corporativo")
        sheet4.cell(row=row, column=10, value="Pendiente Enviar")
        sheet4.cell(row=row, column=11, value="")
        sheet4.cell(row=row, column=12, value="")

    wb.save(excel_path)
    print(f"Se agregaron 5 constructoras al Excel (Filas en Hoja 3: {max_row_3 + 1} a {max_row_3 + len(nuevos)}).")

if __name__ == "__main__":
    append_batch3()
