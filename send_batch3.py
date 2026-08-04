import os.path
import base64
import openpyxl
import time
import requests
import httplib2
from google_auth_httplib2 import AuthorizedHttp
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

requests.packages.urllib3.disable_warnings()
old_request = requests.Session.request
def new_request(*args, **kwargs):
    kwargs['verify'] = False
    return old_request(*args, **kwargs)
requests.Session.request = new_request

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def get_gmail_service():
    creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    http = httplib2.Http(disable_ssl_certificate_validation=True)
    authed_http = AuthorizedHttp(creds, http=http)
    service = build('gmail', 'v1', http=authed_http)
    return service

def create_message(sender, to, subject, message_text):
    message = MIMEMultipart()
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject
    msg = MIMEText(message_text, 'plain')
    message.attach(msg)
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    return {'raw': raw_message}

def send_message(service, user_id, message):
    return (service.users().messages().send(userId=user_id, body=message).execute())

def main():
    service = get_gmail_service()
    excel_path = "docs/Inmobiliarias_Interesadas_Peñablanca_PARDE.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb["3. Correos Personalizados"]
    
    print("Enviando correos al Nuevo Lote de Constructoras (Filas 41-45)...")
    
    # Las 5 nuevas entidades están en las filas 41 a 45
    for row in range(41, 46):
        empresa = sheet.cell(row=row, column=3).value
        correo_1 = str(sheet.cell(row=row, column=4).value or "").strip()
        correo_2 = str(sheet.cell(row=row, column=5).value or "").strip()
        correo_3 = str(sheet.cell(row=row, column=6).value or "").strip()
        asunto = sheet.cell(row=row, column=9).value
        mensaje_cuerpo = sheet.cell(row=row, column=10).value
        
        correos_validos = [e for e in [correo_1, correo_2, correo_3] if '@' in e]
        if not correos_validos or not mensaje_cuerpo:
            continue
            
        to_address = ', '.join(correos_validos)

        try:
            msg = create_message("me", to_address, asunto, mensaje_cuerpo)
            send_message(service, "me", msg)
            print(f"[EXITO] Correo enviado a: {empresa} ({to_address})")
            
            # Actualizar estado en Pipeline 
            sheet4 = wb["4. Pipeline Seguimiento"]
            for r4 in range(5, sheet4.max_row + 1):
                if sheet4.cell(row=r4, column=2).value == empresa:
                    sheet4.cell(row=r4, column=10).value = "Enviado Correo Personalizado" # Columna J
                    break

            time.sleep(2)
        except Exception as e:
            print(f"[ERROR] Falló el envío a {empresa}: {e}")
            
    wb.save(excel_path)
    print("Proceso completado.")

if __name__ == '__main__':
    main()
